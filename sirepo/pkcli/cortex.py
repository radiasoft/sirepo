"""utilities for cortex

:copyright: Copyright (c) 2025 RadiaSoft LLC.  All Rights Reserved.
:license: http://www.apache.org/licenses/LICENSE-2.0.html
"""

from datetime import datetime
from pykern.pkcollections import PKDict
from pykern.pkdebug import pkdc, pkdlog, pkdp
from sqlalchemy.schema import MetaData
import csv
import io
import openmc.data
import openpyxl
import pykern.pkio
import pykern.pkjson
import pykern.sql_db
import re
import requests
import sirepo.auth
import sirepo.quest
import sirepo.sim_api.cortex
import sqlalchemy

_CENTURY = 100.0 * 365 * 24 * 60 * 60


def create_reference_xlsx(xlsx_template_filename, compendium_filename):

    _COMPONENT_ROW = 18

    # This is for the mapping for the default template XLSX file in package_data lib
    _COMMENT_ROWS = PKDict(
        density_comments=7,
        composition_comments=13,
    )
    _FIELD_TO_LOCATION = PKDict(
        {
            "Metadata and Settings": PKDict(
                name="B3",
                material_type="B4",
                availability_factor="B12",
                homogenized_hcpb="B19",
            ),
            "Composition": PKDict(
                url="B3",
                source="B4",
                pointer="B5",
                density_g_cc="B6",
                density_comments=f"B{_COMMENT_ROWS.density_comments}",
                composition_comments=f"B{_COMMENT_ROWS.composition_comments}",
            ),
        }
    )

    def _classify_references(references):
        d = PKDict(url="", pointer="", density_comments="", composition_comments="")
        for r in references:
            # fix missing URL prefix
            for m in re.findall(r"\s\(?(\w+\.\w+\.\w\S*?\/\w\S*?)", r):
                r = r.replace(m, f"https://{m}")
            if "density" in r.lower() or re.search(r"^http\S+$", r):
                m = re.search(r"(http.*?)(?:\s|$)", r, re.IGNORECASE)
                if m and not d.url:
                    d.url = re.sub(r"[.),]*$", "", str(m.group(1)))
                    # only get pointer from first reference
                    m = re.search(
                        r"\b(table|page|figure)\s+(\d.*?),?\s", r, re.IGNORECASE
                    )
                    if m:
                        d.pointer = f"{m.group(1)[0].upper()}{m.group(2)}"
                d.density_comments += (" " if d.density_comments else "") + r
            elif "isotop" in r.lower():
                d.composition_comments += (" " if d.composition_comments else "") + r
        return d

    def _comment(json_data):
        if "Comment" in json_data:
            if isinstance(json_data.Comment, list):
                return json_data.Comment
            return [json_data.Comment]
        return []

    def _create_xlsx(values):
        pkdlog("creating: {}", values.name)
        wb = openpyxl.load_workbook(xlsx_template_filename)
        for n in _FIELD_TO_LOCATION:
            ws = wb[n]
            for f, c in _FIELD_TO_LOCATION[n].items():
                ws[c] = values[f]
        for f, r in _COMMENT_ROWS.items():
            if values[f]:
                # increase cell height if comments are present
                wb["Composition"].row_dimensions[r].height = 100
        for idx in range(len(values.nuclides)):
            c = values.nuclides[idx]
            row = idx + _COMPONENT_ROW
            ws[f"A{row}"] = re.sub(r"-", "", c.isotope)
            ws[f"E{row}"] = c.atom_fraction
        n = re.sub(r"\s+", "_", re.sub(r"[^a-zA-Z0-9 ]", " ", values.name).strip())
        wb.save(f"{n}.xlsx")

    for d in pykern.pkjson.load_any(pykern.pkio.read_text(compendium_filename)).data:
        v = PKDict(
            availability_factor="",
            density_g_cc=d.Density,
            homogenized_hcpb="NO",
            material_type="plasma-facing",
            name=d.Name,
            nuclides=[],
            source="",
        ).pkupdate(_classify_references(_comment(d) + d.References))
        for e in d.Elements:
            for i in e.Isotopes:
                v.nuclides.append(
                    PKDict(
                        isotope=i.Isotope,
                        atom_fraction=i.AtomFraction_whole * 100,
                    )
                )
        _create_xlsx(v)


def gen_components():

    def _livechart_csv():
        r = requests.get(
            "https://nds.iaea.org/relnsd/v1/data?fields=ground_states&nuclides=all"
        )
        r.raise_for_status()
        return io.StringIO(r.text)

    def _parse():
        rv = PKDict(elements=set(), nuclides=set())
        first = True
        for r in csv.reader(_livechart_csv()):
            if not r:
                continue
            if first:
                first = False
                continue
            # num protons
            z = int(r[0])
            if z <= 0:
                continue
            symbol = r[2]
            rv.elements.add(symbol)
            # half_life or from stephen:
            # "Actually, there may be some short-lived lithium isotopes that may be embedded in 1st wall materials, so let’s include all the Lithiums"
            if r[12] != "STABLE" and symbol != "Li":
                # half_life_sec
                if r[16] in ("", "?") or float(r[16]) < _CENTURY:
                    continue
            rv.nuclides.add(r[2] + str(int(r[1]) + z))
        # additional nuclides needed by PNNL examples
        for x in ("Pu238", "Pu241", "Ta180"):
            rv.nuclides.add(x)
        return rv

    def _to_str(components):
        rv = "_COMPONENTS = PKDict(\n"
        i = " " * 4
        j = i * 2
        for c in sorted(components.keys()):
            rv += i + c + "={\n"
            for n in sorted(components[c]):
                rv += f'{j}"{n}",\n'
            rv += i + "},\n"
        return rv + ")\n"

    return _to_str(_parse())


# TODO(robnagler) move to material_db and use pprint
def export_tea(db_file):
    """Returns a python-compatible representation of all materials in
    the specified database."""
    uri = pykern.sql_db.sqlite_uri(pykern.pkio.py_path(db_file))
    m = _populate_materials(_dump_sqlalchemy(uri))
    return f"""# Generated on {datetime.now().strftime("%Y-%m-%dT%H:%M:%SZ")}
MATERIALS = {_json_to_python(pykern.pkjson.dump_pretty(m))}"""


def import_xlsx(uid_or_email, *args):
    def _import_xlsx(filename):
        pkdlog("import: {}", filename)
        p = sirepo.sim_api.cortex.material_xlsx.Parser(filename)
        if p.errors:
            raise AssertionError(
                f"XLSX import {filename} failed with errors: {p.errors}"
            )
        sirepo.sim_api.cortex.material_db.insert_material(parsed=p.result, uid=u)

    if not len(args):
        pykern.pkcli.command_error("missing xlsx file names")
    with sirepo.quest.start() as qcall:
        u = qcall.auth.unchecked_get_user(uid_or_email)
        if not u:
            pykern.pkcli.command_error(f"invalid uid_or_email: {uid_or_email}")
        with qcall.auth.logged_in_user_set(u, method=sirepo.auth.METHOD_EMAIL):
            for f in args:
                _import_xlsx(f)


def _convert_ao_to_wo(materials):

    def _ao_to_wo(ao):
        weight_sum = 0
        weight = PKDict()
        for e in ao:
            if re.search(r"\d", e):
                # Nuclide
                try:
                    w = openmc.data.atomic_mass(e)
                except KeyError as err:
                    raise ValueError(f"Unknown nuclide: {e}")
            else:
                # Element
                if e not in openmc.data.ATOMIC_NUMBER:
                    raise ValueError(f"Unknown element: {e}")
                # may raise ValueError: No naturally-occuring isotopes for element
                w = openmc.data.atomic_weight(e)
            weight[e] = ao[e].target_pct * w / openmc.data.AVOGADRO
            weight_sum += weight[e]

        wo = PKDict()
        for e in ao:
            wo[e] = PKDict(
                target_pct=100.0 * weight[e] / weight_sum,
                min_pct=None,
                max_pct=None,
            )
            if ao[e].max_pct is not None:
                if ao[e].min_pct is None:
                    wo[e].max_pct = wo[e].target_pct
                else:
                    wo[e].min_pct = ao[e].min_pct * wo[e].target_pct / ao[e].target_pct
                    wo[e].max_pct = ao[e].max_pct * wo[e].target_pct / ao[e].target_pct
        return wo

    for m in materials.values():
        if m.is_atom_pct:
            m.components = _ao_to_wo(m.components)
        del m["is_atom_pct"]
    return materials


def _dump_sqlalchemy(uri):
    _DUMP = PKDict(
        material=set(
            [
                "material_id",
                "material_name",
                "is_plasma_facing",
                "structure",
                "microstructure",
                "processing_steps",
                "density_g_cm3",
                "is_atom_pct",
            ]
        ),
        material_component=None,
    )
    e = sqlalchemy.create_engine(uri)
    meta = MetaData()
    meta.reflect(bind=e)
    res = PKDict()
    with e.connect() as conn:
        for t in meta.sorted_tables:
            if t.name not in _DUMP:
                continue
            res[t.name] = [
                PKDict(r._asdict()) for r in conn.execute(sqlalchemy.select(t))
            ]
            if _DUMP[t.name]:
                for r in res[t.name]:
                    for c in t.columns:
                        if c.name not in _DUMP[t.name]:
                            del r[c.name]
    return res


def _json_to_python(value):
    return re.sub(
        r"\bnull\b",
        "None",
        re.sub(r"\bfalse\b", "False", re.sub(r"\btrue\b", "True", value)),
    )


def _populate_materials(dump):
    res = PKDict()
    for m in dump.material:
        res[m.material_name] = m
        m.components = PKDict()
        for c in dump.material_component:
            if c.material_id == m.material_id:
                mc = c.copy()
                for n in (
                    "material_component_id",
                    "material_component_name",
                    "material_id",
                ):
                    del mc[n]
                m.components[c.material_component_name.capitalize()] = mc
        for n in ("material_id", "material_name"):
            del m[n]
    return _convert_ao_to_wo(res)
